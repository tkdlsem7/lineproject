# backend/ScheduleHub/routers.py
from __future__ import annotations
from openpyxl.utils.datetime import from_excel

import re
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import func, or_, text
from sqlalchemy.orm import Session

from backend.deps import get_db
from .models import (
    EquipmentMaster,
    ImportRawRow,
    ScheduleEvent,
    ScheduleEventHistory,
    UploadFile as UploadFileModel,
)

router = APIRouter(
    prefix="/schedule-hub",
    tags=["Schedule Hub"],
)

UPLOAD_DIR = Path(__file__).resolve().parent / "uploaded_schedule_files"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

TEAM_SOURCE_MAP = {
    "생산일정": "production",
    "출하일정": "shipment",
    "개조": "remodel",
    "인터페이스": "interface",
    "MANI": "mani",
    "OPUS": "opus",
    "칠러": "chiller",
    "제조일정": "manufacturing",
    "production": "production",
    "shipment": "shipment",
    "remodel": "remodel",
    "interface": "interface",
    "mani": "mani",
    "opus": "opus",
    "chiller": "chiller",
    "manufacturing": "manufacturing",
}


# ------------------------------------------------------------------
# 공통 유틸
# ------------------------------------------------------------------
def _normalize_team_name(team_name: str) -> str:
    t = (team_name or "").strip()
    if t not in TEAM_SOURCE_MAP:
        raise HTTPException(
            status_code=400,
            detail="지원하지 않는 team_name 입니다. (생산일정/출하일정/개조/인터페이스/MANI/OPUS/칠러/제조일정)",
        )
    return TEAM_SOURCE_MAP[t]


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = s.replace("\n", "").replace("\r", "").replace(" ", "").replace("_", "")
    return s


def _text(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _normalize_model_text(value: Any) -> Optional[str]:
    s = _text(value)
    if not s:
        return None
    s = re.sub(r"\s+", "", s)
    return s or None


def _normalize_batch_text(value: Any) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, float):
        if value.is_integer():
            s = str(int(value))
        else:
            s = str(value).strip()
    else:
        s = str(value).strip()

    if not s or s == "-":
        return None

    if re.fullmatch(r"\d+", s):
        return s.zfill(2)

    return s


def _parse_positive_int(value: Any) -> Optional[int]:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, int):
        return value if value > 0 else None

    if isinstance(value, float):
        if value <= 0:
            return None
        return int(value)

    s = str(value).strip().replace(",", "")
    if not s or s == "-":
        return None

    if re.fullmatch(r"\d+", s):
        qty = int(s)
        return qty if qty > 0 else None

    m = re.fullmatch(r"(\d+)\s*(ea|EA|대)?", s)
    if m:
        qty = int(m.group(1))
        return qty if qty > 0 else None

    return None


def _jsonable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def _to_date(value: Any) -> Optional[date]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    # 엑셀 숫자 날짜 대응
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None

        m = re.search(r"(20\d{2})[-./](\d{1,2})[-./](\d{1,2})", s)
        if m:
            y, mm, dd = map(int, m.groups())
            try:
                return date(y, mm, dd)
            except ValueError:
                return None

        m = re.search(r"\b(\d{2})[-./](\d{1,2})[-./](\d{1,2})\b", s)
        if m:
            yy, mm, dd = map(int, m.groups())
            y = 2000 + yy
            try:
                return date(y, mm, dd)
            except ValueError:
                return None

        return None

    return None


def _looks_like_machine_no(value: Any) -> Optional[str]:
    s = _text(value)
    if not s:
        return None

    s = s.replace(" ", "")

    exact_bad_words = {
        "칠러만출하",
        "품질TEST용",
        "사외창고보관중",
        "DEMO반입",
        "출하완료",
        "-",
    }
    contains_bad_words = [
        "반입",
    ]

    if s in exact_bad_words:
        return None

    if any(word in s for word in contains_bad_words):
        return None

    if re.fullmatch(r"\d+(\.\d+)?", s):
        return None

    if "-" not in s:
        return None

    return s


def _pick_machine_no(raw_data: dict[str, Any], *values: Any) -> Optional[str]:
    for value in values:
        machine_no = _looks_like_machine_no(value)
        if machine_no:
            return machine_no

    for key in ("호기", "장비호기", "설비번호", "위치", "machine", "장 비"):
        machine_no = _looks_like_machine_no(raw_data.get(key))
        if machine_no:
            return machine_no

    return None


def _find_header_row(ws, required_keywords: list[str], scan_rows: int = 15, scan_cols: int = 80) -> Optional[int]:
    best_row = None
    best_score = -1

    normalized_required = [_norm_header(x) for x in required_keywords]

    for r in range(1, min(ws.max_row, scan_rows) + 1):
        row_values = []
        for c in range(1, min(ws.max_column, scan_cols) + 1):
            v = ws.cell(r, c).value
            nv = _norm_header(v)
            if nv:
                row_values.append(nv)

        if not row_values:
            continue

        score = 0
        for req in normalized_required:
            if any(req in cell or cell in req for cell in row_values):
                score += 1

        if score > best_score:
            best_score = score
            best_row = r

    if best_score < max(2, len(required_keywords) // 2):
        return None

    return best_row


def _prepare_sheet(ws, required_keywords: list[str]):
    header_row = _find_header_row(ws, required_keywords)
    if not header_row:
        return None

    header_map: dict[str, int] = {}
    headers: list[tuple[int, str]] = []
    last_col = 0

    for c in range(1, min(ws.max_column, 100) + 1):
        raw = ws.cell(header_row, c).value
        text = _text(raw)
        if not text:
            continue
        headers.append((c, text))
        header_map[_norm_header(text)] = c
        last_col = c

    if not headers:
        return None

    return header_row, header_map, headers, last_col


def _find_col(header_map: dict[str, int], *candidates: str) -> Optional[int]:
    norm_candidates = [_norm_header(x) for x in candidates]

    for cand in norm_candidates:
        for h, col in header_map.items():
            if cand == h or cand in h or h in cand:
                return col
    return None


def _collect_raw_data(ws, row_idx: int, headers: list[tuple[int, str]]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for col, header_text in headers:
        value = ws.cell(row_idx, col).value
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        raw[header_text] = _jsonable(value)
    return raw


def _save_raw_row(
    db: Session,
    upload_file_id: int,
    sheet_name: str,
    row_no: int,
    machine_no: Optional[str],
    raw_data: dict[str, Any],
    parse_status: str = "parsed",
    error_message: Optional[str] = None,
):
    db.add(
        ImportRawRow(
            upload_file_id=upload_file_id,
            sheet_name=sheet_name,
            row_no=row_no,
            machine_no=machine_no,
            raw_data=raw_data,
            parse_status=parse_status,
            error_message=error_message,
        )
    )


def _build_generated_machine_no(model_text: str, batch_text: str, unit_no: int) -> str:
    return f"{model_text}-{batch_text}-{int(unit_no):02d}"


def _upsert_equipment(
    db: Session,
    *,
    machine_no: str,
    model: Optional[str] = None,
    customer_name: Optional[str] = None,
    stage_sn: Optional[str] = None,
    loader_sn: Optional[str] = None,
    cold_type: Optional[str] = None,
    mani_type: Optional[str] = None,
    current_status: Optional[str] = None,
    is_shipped: Optional[bool] = None,
) -> EquipmentMaster:
    row = (
        db.query(EquipmentMaster)
        .filter(EquipmentMaster.machine_no == machine_no)
        .first()
    )

    if not row:
        row = EquipmentMaster(machine_no=machine_no)
        db.add(row)
        db.flush()

    if model:
        row.model = model
    if customer_name:
        row.customer_name = customer_name
    if stage_sn:
        row.stage_sn = stage_sn
    if loader_sn:
        row.loader_sn = loader_sn
    if cold_type is not None and cold_type != "":
        row.cold_type = cold_type
    if mani_type is not None and mani_type != "":
        row.mani_type = mani_type
    if current_status:
        row.current_status = current_status
    if is_shipped is not None:
        row.is_shipped = bool(is_shipped)

    row.updated_at = datetime.now()
    db.flush()
    return row


def _add_event(
    db: Session,
    seen: set[tuple],
    *,
    equipment_id: int,
    source_type: str,
    event_type: str,
    event_name: str,
    event_date: Optional[date],
    status: Optional[str],
    team_name: str,
    mo_no: Optional[str],
    extra_data: Optional[dict[str, Any]],
    source_file_id: int,
    source_sheet_name: str,
    source_row_no: int,
) -> int:
    if not event_date:
        return 0

    key = (equipment_id, source_type, event_type, mo_no or "")
    if key in seen:
        return 0
    seen.add(key)

    db.add(
        ScheduleEvent(
            equipment_id=equipment_id,
            source_type=source_type,
            event_type=event_type,
            event_name=event_name,
            event_date=event_date,
            status=status,
            team_name=team_name,
            mo_no=mo_no,
            extra_data=extra_data or {},
            source_file_id=source_file_id,
            source_sheet_name=source_sheet_name,
            source_row_no=source_row_no,
            updated_at=datetime.now(),
        )
    )
    return 1



def _event_identity_key(row: ScheduleEvent) -> tuple[int, str, str, str]:
    return (
        int(row.equipment_id),
        str(row.source_type or ""),
        str(row.event_type or ""),
        str(row.mo_no or ""),
    )


def _normalize_compare_extra(extra: Any) -> dict[str, Any]:
    if not isinstance(extra, dict):
        return {}

    normalized: dict[str, Any] = {}
    for key, value in sorted(extra.items(), key=lambda item: str(item[0])):
        if value in (None, "", [], {}):
            continue
        normalized[str(key)] = value
    return normalized


def _events_are_same(before: ScheduleEvent, after: ScheduleEvent) -> bool:
    return (
        before.event_date == after.event_date
        and (before.status or None) == (after.status or None)
        and (before.event_name or "") == (after.event_name or "")
        and (before.team_name or None) == (after.team_name or None)
        and _normalize_compare_extra(before.extra_data) == _normalize_compare_extra(after.extra_data)
    )


def _extract_change_reason(*extras: Any) -> Optional[str]:
    candidate_keys = [
        "change_reason",
        "changed_text",
        "delay_reason",
        "reason",
        "변경사유",
        "지연사유",
        "변경 유무",
        "변경유무",
        "지연 여부",
        "지연여부",
        "비고",
        "remark",
        "remarks",
        "note",
    ]

    for extra in extras:
        if not isinstance(extra, dict):
            continue
        for key in candidate_keys:
            value = extra.get(key)
            text_value = _text(value)
            if text_value:
                return text_value
    return None


def _build_event_map(rows: list[ScheduleEvent]) -> dict[tuple[int, str, str, str], ScheduleEvent]:
    event_map: dict[tuple[int, str, str, str], ScheduleEvent] = {}
    for row in rows:
        event_map[_event_identity_key(row)] = row
    return event_map


def _create_schedule_event_history(
    db: Session,
    *,
    before: Optional[ScheduleEvent],
    after: Optional[ScheduleEvent],
    changed_by: Optional[str],
) -> None:
    if not before and not after:
        return

    change_type = "updated"
    if before and not after:
        change_type = "deleted"
    elif after and not before:
        change_type = "inserted"

    db.add(
        ScheduleEventHistory(
            equipment_id=(after.equipment_id if after else before.equipment_id),
            source_type=(after.source_type if after else before.source_type),
            event_type=(after.event_type if after else before.event_type),
            event_name=(after.event_name if after else before.event_name),
            team_name=(after.team_name if after else before.team_name),
            mo_no=(after.mo_no if after else before.mo_no),
            change_type=change_type,
            before_event_date=(before.event_date if before else None),
            before_status=(before.status if before else None),
            before_extra_data=(_normalize_compare_extra(before.extra_data) if before else {}),
            before_source_file_id=(before.source_file_id if before else None),
            before_source_sheet_name=(before.source_sheet_name if before else None),
            before_source_row_no=(before.source_row_no if before else None),
            after_event_date=(after.event_date if after else None),
            after_status=(after.status if after else None),
            after_extra_data=(_normalize_compare_extra(after.extra_data) if after else {}),
            after_source_file_id=(after.source_file_id if after else None),
            after_source_sheet_name=(after.source_sheet_name if after else None),
            after_source_row_no=(after.source_row_no if after else None),
            changed_by=_text(changed_by),
            change_reason=_extract_change_reason(
                after.extra_data if after else None,
                before.extra_data if before else None,
            ),
            created_at=datetime.now(),
        )
    )


def _record_schedule_event_history(
    db: Session,
    *,
    source_type: str,
    new_source_file_id: int,
    changed_by: Optional[str],
) -> tuple[int, int, int]:
    previous_rows = (
        db.query(ScheduleEvent)
        .filter(
            ScheduleEvent.source_type == source_type,
            ScheduleEvent.source_file_id != new_source_file_id,
        )
        .order_by(ScheduleEvent.id.asc())
        .all()
    )

    if not previous_rows:
        return 0, 0, 0

    current_rows = (
        db.query(ScheduleEvent)
        .filter(
            ScheduleEvent.source_type == source_type,
            ScheduleEvent.source_file_id == new_source_file_id,
        )
        .order_by(ScheduleEvent.id.asc())
        .all()
    )

    previous_map = _build_event_map(previous_rows)
    current_map = _build_event_map(current_rows)

    inserted_count = 0
    updated_count = 0
    deleted_count = 0

    all_keys = set(previous_map.keys()) | set(current_map.keys())
    for key in sorted(all_keys):
        before = previous_map.get(key)
        after = current_map.get(key)

        if before and after:
            if _events_are_same(before, after):
                continue
            _create_schedule_event_history(db, before=before, after=after, changed_by=changed_by)
            updated_count += 1
            continue

        if before and not after:
            _create_schedule_event_history(db, before=before, after=None, changed_by=changed_by)
            deleted_count += 1
            continue

        if after and not before:
            _create_schedule_event_history(db, before=None, after=after, changed_by=changed_by)
            inserted_count += 1

    return inserted_count, updated_count, deleted_count


# ------------------------------------------------------------------
# 팀별 파서
# ------------------------------------------------------------------
def _process_production_workbook(db: Session, wb, upload_row: UploadFileModel, team_name: str) -> tuple[int, int]:
    source_type = "production"
    seen: set[tuple] = set()
    parsed_rows = 0
    event_count = 0

    for ws in wb.worksheets:
        prepared = _prepare_sheet(ws, ["호기", "고객", "출하요청일", "setting시작일"])
        if not prepared:
            continue

        header_row, header_map, headers, _ = prepared

        machine_col = _find_col(header_map, "호기", "장비호기", "설비번호", "위치", "machine")
        customer_col = _find_col(header_map, "고객", "customer")
        cold_col = _find_col(header_map, "cold")
        mani_col = _find_col(header_map, "mani")
        mo_col = _find_col(header_map, "제작넘버", "mo번호")
        model_col = _find_col(header_map, "장비사양", "모델")
        status_col = _find_col(header_map, "장비상태", "장비 상태")
        stage_sn_col = _find_col(header_map, "sts/n", "stsn", "st s/n")
        loader_sn_col = _find_col(header_map, "lds/n", "ldsn", "ld s/n")
        ship_col = _find_col(header_map, "출하요청일")
        setting_start_col = _find_col(header_map, "setting시작일", "setting 시작일")
        setting_end_col = _find_col(header_map, "setting종료일", "setting 종료일")
        qc_start_col = _find_col(header_map, "qc시작", "qc 시작")
        qc_end_col = _find_col(header_map, "qc종료", "qc 종료")
        changed_col = _find_col(header_map, "변경유무")
        prev_ship_col = _find_col(header_map, "출하일변경전", "출하일 변경전")

        if not machine_col:
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_data = _collect_raw_data(ws, row_idx, headers)
            if not raw_data:
                continue

            parsed_rows += 1

            try:
                machine_no = _pick_machine_no(
                    raw_data,
                    ws.cell(row_idx, machine_col).value if machine_col else None,
                )

                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    machine_no,
                    raw_data,
                    "parsed",
                    None,
                )

                if not machine_no:
                    continue

                current_status = _text(ws.cell(row_idx, status_col).value) if status_col else None
                is_shipped = None
                if current_status:
                    is_shipped = "출하" in current_status

                equipment = _upsert_equipment(
                    db,
                    machine_no=machine_no,
                    model=_text(ws.cell(row_idx, model_col).value) if model_col else None,
                    customer_name=_text(ws.cell(row_idx, customer_col).value) if customer_col else None,
                    stage_sn=_text(ws.cell(row_idx, stage_sn_col).value) if stage_sn_col else None,
                    loader_sn=_text(ws.cell(row_idx, loader_sn_col).value) if loader_sn_col else None,
                    cold_type=_text(ws.cell(row_idx, cold_col).value) if cold_col else None,
                    mani_type=_text(ws.cell(row_idx, mani_col).value) if mani_col else None,
                    current_status=current_status,
                    is_shipped=is_shipped,
                )

                mo_no = _text(ws.cell(row_idx, mo_col).value) if mo_col else None
                ship_date = _to_date(ws.cell(row_idx, ship_col).value) if ship_col else None
                prev_ship_date = _to_date(ws.cell(row_idx, prev_ship_col).value) if prev_ship_col else None
                changed_text = _text(ws.cell(row_idx, changed_col).value) if changed_col else None

                shipment_extra: dict[str, Any] = {}
                shipment_status: Optional[str] = None
                if prev_ship_date and ship_date and prev_ship_date != ship_date:
                    shipment_status = "변경"
                    shipment_extra["previous_date"] = prev_ship_date.isoformat()
                elif changed_text:
                    shipment_status = "변경"
                    shipment_extra["changed_text"] = changed_text

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="shipment_request",
                    event_name="출하요청",
                    event_date=ship_date,
                    status=shipment_status,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data=shipment_extra,
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="setting_start",
                    event_name="SETTING 시작",
                    event_date=_to_date(ws.cell(row_idx, setting_start_col).value) if setting_start_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="setting_end",
                    event_name="SETTING 종료",
                    event_date=_to_date(ws.cell(row_idx, setting_end_col).value) if setting_end_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="qc_start",
                    event_name="QC 시작",
                    event_date=_to_date(ws.cell(row_idx, qc_start_col).value) if qc_start_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="qc_end",
                    event_name="QC 종료",
                    event_date=_to_date(ws.cell(row_idx, qc_end_col).value) if qc_end_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

            except Exception as e:
                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    None,
                    raw_data,
                    "error",
                    str(e)[:500],
                )

    return parsed_rows, event_count


def _process_shipment_workbook(db: Session, wb, upload_row: UploadFileModel, team_name: str) -> tuple[int, int]:
    source_type = "shipment"
    seen: set[tuple] = set()
    parsed_rows = 0
    event_count = 0

    for ws in wb.worksheets:
        prepared = _prepare_sheet(ws, ["호기", "출하요청일", "제작넘버"])
        if not prepared:
            continue

        header_row, header_map, headers, _ = prepared

        machine_col = _find_col(header_map, "호기", "장비호기", "설비번호", "위치", "machine")
        customer_col = _find_col(header_map, "customer", "고객")
        cold_col = _find_col(header_map, "cold")
        mani_col = _find_col(header_map, "mani")
        mo_col = _find_col(header_map, "제작넘버")
        model_col = _find_col(header_map, "장비사양", "모델")
        stage_sn_col = _find_col(header_map, "stages/n", "stagesn", "s/n", "st s/n")
        loader_sn_col = _find_col(header_map, "loaders/n", "loadersn", "ld s/n")
        ship_col = _find_col(header_map, "출하요청일")
        plan_col = _find_col(header_map, "판매계획")
        material_in_col = _find_col(header_map, "자재입고일")
        inhouse_in_col = _find_col(header_map, "사내장비입고일")
        electric_done_col = _find_col(header_map, "전장완료일")
        ship_ready_col = _find_col(header_map, "출하가능일")

        if not machine_col:
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_data = _collect_raw_data(ws, row_idx, headers)
            if not raw_data:
                continue

            parsed_rows += 1

            try:
                machine_no = _pick_machine_no(
                    raw_data,
                    ws.cell(row_idx, machine_col).value if machine_col else None,
                )

                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    machine_no,
                    raw_data,
                    "parsed",
                    None,
                )

                if not machine_no:
                    continue

                current_status = _text(ws.cell(row_idx, plan_col).value) if plan_col else None

                equipment = _upsert_equipment(
                    db,
                    machine_no=machine_no,
                    model=_text(ws.cell(row_idx, model_col).value) if model_col else None,
                    customer_name=_text(ws.cell(row_idx, customer_col).value) if customer_col else None,
                    stage_sn=_text(ws.cell(row_idx, stage_sn_col).value) if stage_sn_col else None,
                    loader_sn=_text(ws.cell(row_idx, loader_sn_col).value) if loader_sn_col else None,
                    cold_type=_text(ws.cell(row_idx, cold_col).value) if cold_col else None,
                    mani_type=_text(ws.cell(row_idx, mani_col).value) if mani_col else None,
                    current_status=current_status,
                    is_shipped=("출하" in current_status) if current_status else None,
                )

                mo_no = _text(ws.cell(row_idx, mo_col).value) if mo_col else None

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="shipment_request",
                    event_name="출하요청",
                    event_date=_to_date(ws.cell(row_idx, ship_col).value) if ship_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="material_in",
                    event_name="자재 입고",
                    event_date=_to_date(ws.cell(row_idx, material_in_col).value) if material_in_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="equipment_in",
                    event_name="사내장비 입고",
                    event_date=_to_date(ws.cell(row_idx, inhouse_in_col).value) if inhouse_in_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="electric_done",
                    event_name="전장 완료",
                    event_date=_to_date(ws.cell(row_idx, electric_done_col).value) if electric_done_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="shipment_ready",
                    event_name="출하 가능",
                    event_date=_to_date(ws.cell(row_idx, ship_ready_col).value) if ship_ready_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

            except Exception as e:
                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    None,
                    raw_data,
                    "error",
                    str(e)[:500],
                )

    return parsed_rows, event_count


def _process_remodel_workbook(db: Session, wb, upload_row: UploadFileModel, team_name: str) -> tuple[int, int]:
    source_type = "remodel"
    seen: set[tuple] = set()
    parsed_rows = 0
    event_count = 0

    for ws in wb.worksheets:
        prepared = _prepare_sheet(ws, ["장비", "고객사", "개조일정"])
        if not prepared:
            continue

        header_row, header_map, headers, _ = prepared

        machine_col = _find_col(header_map, "장비", "호기", "장비호기", "설비번호")
        customer_col = _find_col(header_map, "고객사")
        model_col = _find_col(header_map, "모델")
        mo_col = _find_col(header_map, "mo번호")
        status_col = _find_col(header_map, "상태")
        material_in_col = _find_col(header_map, "자재입고일")
        remodel_date_col = _find_col(header_map, "개조일정")

        if not machine_col:
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_data = _collect_raw_data(ws, row_idx, headers)
            if not raw_data:
                continue

            parsed_rows += 1

            try:
                machine_no = _pick_machine_no(
                    raw_data,
                    ws.cell(row_idx, machine_col).value if machine_col else None,
                )

                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    machine_no,
                    raw_data,
                    "parsed",
                    None,
                )

                if not machine_no:
                    continue

                equipment = _upsert_equipment(
                    db,
                    machine_no=machine_no,
                    model=_text(ws.cell(row_idx, model_col).value) if model_col else None,
                    customer_name=_text(ws.cell(row_idx, customer_col).value) if customer_col else None,
                    current_status=_text(ws.cell(row_idx, status_col).value) if status_col else None,
                )

                mo_no = _text(ws.cell(row_idx, mo_col).value) if mo_col else None

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="remodel_material_in",
                    event_name="개조 자재입고",
                    event_date=_to_date(ws.cell(row_idx, material_in_col).value) if material_in_col else None,
                    status="변경" if raw_data.get("지연 여부") else None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="remodel_schedule",
                    event_name="개조 일정",
                    event_date=_to_date(ws.cell(row_idx, remodel_date_col).value) if remodel_date_col else None,
                    status="변경" if raw_data.get("지연 여부") else None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

            except Exception as e:
                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    None,
                    raw_data,
                    "error",
                    str(e)[:500],
                )

    return parsed_rows, event_count


def _process_interface_workbook(db: Session, wb, upload_row: UploadFileModel, team_name: str) -> tuple[int, int]:
    source_type = "interface"
    seen: set[tuple] = set()
    parsed_rows = 0
    event_count = 0

    for ws in wb.worksheets:
        prepared = _prepare_sheet(ws, ["호기", "제작넘버", "인터페이스입고예정일"])
        if not prepared:
            continue

        header_row, header_map, headers, _ = prepared

        machine_col = _find_col(header_map, "호기", "장비호기", "설비번호", "위치", "machine")
        customer_col = _find_col(header_map, "customer", "고객")
        cold_col = _find_col(header_map, "cold")
        mani_col = _find_col(header_map, "mani")
        mo_col = _find_col(header_map, "제작넘버")
        model_col = _find_col(header_map, "장비사양", "모델")
        ship_col = _find_col(header_map, "출하요청일")
        order_col = _find_col(header_map, "인터페이스발주일")
        in_exp_col = _find_col(header_map, "인터페이스입고예정일")
        in_done_col = _find_col(header_map, "인터페이스입고완료일")
        release_ready_col = _find_col(header_map, "인터페이스불출가능일")
        release_col = _find_col(header_map, "인터페이스불출일자")

        if not machine_col:
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_data = _collect_raw_data(ws, row_idx, headers)
            if not raw_data:
                continue

            parsed_rows += 1

            try:
                machine_no = _pick_machine_no(
                    raw_data,
                    ws.cell(row_idx, machine_col).value if machine_col else None,
                )

                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    machine_no,
                    raw_data,
                    "parsed",
                    None,
                )

                if not machine_no:
                    continue

                equipment = _upsert_equipment(
                    db,
                    machine_no=machine_no,
                    model=_text(ws.cell(row_idx, model_col).value) if model_col else None,
                    customer_name=_text(ws.cell(row_idx, customer_col).value) if customer_col else None,
                    cold_type=_text(ws.cell(row_idx, cold_col).value) if cold_col else None,
                    mani_type=_text(ws.cell(row_idx, mani_col).value) if mani_col else None,
                )

                mo_no = _text(ws.cell(row_idx, mo_col).value) if mo_col else None
                ship_date = _to_date(ws.cell(row_idx, ship_col).value) if ship_col else None

                if ship_date:
                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="shipment_request",
                        event_name="출하요청",
                        event_date=ship_date,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data={},
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="interface_order",
                    event_name="인터페이스 발주",
                    event_date=_to_date(ws.cell(row_idx, order_col).value) if order_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="interface_in_expected",
                    event_name="인터페이스 입고예정",
                    event_date=_to_date(ws.cell(row_idx, in_exp_col).value) if in_exp_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="interface_in_done",
                    event_name="인터페이스 입고완료",
                    event_date=_to_date(ws.cell(row_idx, in_done_col).value) if in_done_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="interface_release_ready",
                    event_name="인터페이스 불출가능",
                    event_date=_to_date(ws.cell(row_idx, release_ready_col).value) if release_ready_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="interface_release",
                    event_name="인터페이스 불출",
                    event_date=_to_date(ws.cell(row_idx, release_col).value) if release_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

            except Exception as e:
                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    None,
                    raw_data,
                    "error",
                    str(e)[:500],
                )

    return parsed_rows, event_count


def _process_mani_workbook(db: Session, wb, upload_row: UploadFileModel, team_name: str) -> tuple[int, int]:
    source_type = "mani"
    seen: set[tuple] = set()
    parsed_rows = 0
    event_count = 0

    for ws in wb.worksheets:
        prepared = _prepare_sheet(ws, ["호기", "mani입고일", "prober s/n"])
        if not prepared:
            continue

        header_row, header_map, headers, _ = prepared

        machine_col = _find_col(header_map, "호기", "장비호기", "설비번호", "위치", "machine")
        customer_col = _find_col(header_map, "고객사", "customer")
        mo_col = _find_col(header_map, "제작넘버", "제작 넘버", "mo번호")
        material_in_col = _find_col(header_map, "자재입고일", "자재 입고일")
        mani_in_col = _find_col(header_map, "mani입고일", "mani 입고일")
        ship_col = _find_col(header_map, "출하일")
        assembly_start_col = _find_col(header_map, "조립시작일", "조립 시작일")
        assembly_end_col = _find_col(header_map, "조립완료일", "조립 완료일")
        prober_sn_col = _find_col(header_map, "prober s/n", "probers/n", "prober")
        mani_sn_col = _find_col(header_map, "mani s/n", "manis/n", "mani sn")

        if not machine_col:
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_data = _collect_raw_data(ws, row_idx, headers)
            if not raw_data:
                continue

            parsed_rows += 1

            try:
                raw_machine_value = ws.cell(row_idx, machine_col).value if machine_col else None
                machine_no = _pick_machine_no(raw_data, raw_machine_value)

                customer_name = _text(ws.cell(row_idx, customer_col).value) if customer_col else None
                mo_no = _text(ws.cell(row_idx, mo_col).value) if mo_col else None
                prober_sn = _text(ws.cell(row_idx, prober_sn_col).value) if prober_sn_col else None
                mani_sn = _text(ws.cell(row_idx, mani_sn_col).value) if mani_sn_col else None

                equipment = None

                # 1) 일반 케이스: machine_no가 정상 문자열일 때
                if machine_no:
                    equipment = _upsert_equipment(
                        db,
                        machine_no=machine_no,
                        customer_name=customer_name,
                    )

                # 2) MANI 전용 케이스: PROBER S/N -> 기존 장비의 stage_sn / loader_sn 매핑
                if not equipment and prober_sn:
                    prober_sn_norm = re.sub(r"\s+", "", prober_sn).upper()

                    equipment = (
                        db.query(EquipmentMaster)
                        .filter(
                            or_(
                                func.upper(func.replace(EquipmentMaster.stage_sn, " ", "")) == prober_sn_norm,
                                func.upper(func.replace(EquipmentMaster.loader_sn, " ", "")) == prober_sn_norm,
                            )
                        )
                        .first()
                    )

                    if equipment:
                        machine_no = equipment.machine_no

                        # 고객사 비어있을 때만 보정
                        if customer_name and not equipment.customer_name:
                            equipment.customer_name = customer_name

                        equipment.updated_at = datetime.now()
                        db.flush()

                # raw row 저장은 매핑 후 machine_no 기준으로 저장
                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    machine_no,
                    raw_data,
                    "parsed",
                    None,
                )

                if not equipment:
                    continue

                extra: dict[str, Any] = {}
                if prober_sn:
                    extra["prober_sn"] = prober_sn
                if mani_sn:
                    extra["mani_sn"] = mani_sn
                if raw_machine_value is not None:
                    extra["source_hogi"] = _jsonable(raw_machine_value)

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="mani_material_in",
                    event_name="MANI 자재입고",
                    event_date=_to_date(ws.cell(row_idx, material_in_col).value) if material_in_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data=extra,
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="mani_in",
                    event_name="MANI 입고",
                    event_date=_to_date(ws.cell(row_idx, mani_in_col).value) if mani_in_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data=extra,
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="mani_assembly_start",
                    event_name="MANI 조립 시작",
                    event_date=_to_date(ws.cell(row_idx, assembly_start_col).value) if assembly_start_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data=extra,
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="mani_assembly_end",
                    event_name="MANI 조립 완료",
                    event_date=_to_date(ws.cell(row_idx, assembly_end_col).value) if assembly_end_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data=extra,
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="mani_out",
                    event_name="MANI 출하",
                    event_date=_to_date(ws.cell(row_idx, ship_col).value) if ship_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data=extra,
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

            except Exception as e:
                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    None,
                    raw_data,
                    "error",
                    str(e)[:500],
                )

    return parsed_rows, event_count


def _process_opus_workbook(db: Session, wb, upload_row: UploadFileModel, team_name: str) -> tuple[int, int]:
    parsed_rows = 0
    event_count = 0
    target_sheet = None

    for ws in wb.worksheets:
        if "보유장비리스트" in ws.title:
            target_sheet = ws
            break

    if not target_sheet:
        return parsed_rows, event_count

    prepared = _prepare_sheet(target_sheet, ["호기", "제품모델", "판매계획"])
    if not prepared:
        return parsed_rows, event_count

    header_row, header_map, headers, _ = prepared

    machine_col = _find_col(header_map, "호기", "장비호기", "설비번호", "위치", "machine")
    product_model_col = _find_col(header_map, "제품모델")
    model_col = _find_col(header_map, "장비사양", "모델")
    stage_sn_col = _find_col(header_map, "stages/n", "stagesn", "st s/n")
    loader_sn_col = _find_col(header_map, "loaders/n", "loadersn", "ld s/n")
    customer_col = _find_col(header_map, "customer", "고객")
    cold_col = _find_col(header_map, "cold")
    mani_col = _find_col(header_map, "mani")
    plan_col = _find_col(header_map, "판매계획")

    if not machine_col:
        return parsed_rows, event_count

    for row_idx in range(header_row + 1, target_sheet.max_row + 1):
        raw_data = _collect_raw_data(target_sheet, row_idx, headers)
        if not raw_data:
            continue

        parsed_rows += 1

        try:
            machine_no = _pick_machine_no(
                raw_data,
                target_sheet.cell(row_idx, machine_col).value if machine_col else None,
            )

            _save_raw_row(
                db,
                upload_row.id,
                target_sheet.title,
                row_idx,
                machine_no,
                raw_data,
                "parsed",
                None,
            )

            if not machine_no:
                continue

            plan = _text(target_sheet.cell(row_idx, plan_col).value) if plan_col else None

            _upsert_equipment(
                db,
                machine_no=machine_no,
                model=(
                    _text(target_sheet.cell(row_idx, model_col).value) if model_col else None
                ) or (
                    _text(target_sheet.cell(row_idx, product_model_col).value) if product_model_col else None
                ),
                customer_name=_text(target_sheet.cell(row_idx, customer_col).value) if customer_col else None,
                stage_sn=_text(target_sheet.cell(row_idx, stage_sn_col).value) if stage_sn_col else None,
                loader_sn=_text(target_sheet.cell(row_idx, loader_sn_col).value) if loader_sn_col else None,
                cold_type=_text(target_sheet.cell(row_idx, cold_col).value) if cold_col else None,
                mani_type=_text(target_sheet.cell(row_idx, mani_col).value) if mani_col else None,
                current_status=plan,
                is_shipped=False,
            )

        except Exception as e:
            _save_raw_row(
                db,
                upload_row.id,
                target_sheet.title,
                row_idx,
                None,
                raw_data,
                "error",
                str(e)[:500],
            )

    return parsed_rows, event_count


def _process_chiller_workbook(db: Session, wb, upload_row: UploadFileModel, team_name: str) -> tuple[int, int]:
    source_type = "chiller"
    seen: set[tuple] = set()
    parsed_rows = 0
    event_count = 0

    for ws in wb.worksheets:
        prepared = _prepare_sheet(ws, ["장비호기", "입고일", "출하일"])
        if not prepared:
            continue

        header_row, header_map, headers, _ = prepared

        machine_col = _find_col(header_map, "장비호기", "호기", "설비번호", "위치", "machine")
        customer_col = _find_col(header_map, "고객사", "customer")
        model_col = _find_col(header_map, "모델")
        mo_col = _find_col(header_map, "제작의뢰")
        expected_in_col = _find_col(header_map, "입고예정일")
        in_col = _find_col(header_map, "입고일")
        out_col = _find_col(header_map, "출하일")

        if not machine_col:
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_data = _collect_raw_data(ws, row_idx, headers)
            if not raw_data:
                continue

            parsed_rows += 1

            try:
                machine_no = _pick_machine_no(
                    raw_data,
                    ws.cell(row_idx, machine_col).value if machine_col else None,
                )

                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    machine_no,
                    raw_data,
                    "parsed",
                    None,
                )

                if not machine_no:
                    continue

                equipment = _upsert_equipment(
                    db,
                    machine_no=machine_no,
                    customer_name=_text(ws.cell(row_idx, customer_col).value) if customer_col else None,
                )

                mo_no = _text(ws.cell(row_idx, mo_col).value) if mo_col else None

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="chiller_in_expected",
                    event_name="칠러 입고예정",
                    event_date=_to_date(ws.cell(row_idx, expected_in_col).value) if expected_in_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={"model": _text(ws.cell(row_idx, model_col).value) if model_col else None},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="chiller_in",
                    event_name="칠러 입고",
                    event_date=_to_date(ws.cell(row_idx, in_col).value) if in_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={"model": _text(ws.cell(row_idx, model_col).value) if model_col else None},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

                event_count += _add_event(
                    db,
                    seen,
                    equipment_id=equipment.id,
                    source_type=source_type,
                    event_type="chiller_out",
                    event_name="칠러 출하",
                    event_date=_to_date(ws.cell(row_idx, out_col).value) if out_col else None,
                    status=None,
                    team_name=team_name,
                    mo_no=mo_no,
                    extra_data={"model": _text(ws.cell(row_idx, model_col).value) if model_col else None},
                    source_file_id=upload_row.id,
                    source_sheet_name=ws.title,
                    source_row_no=row_idx,
                )

            except Exception as e:
                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    None,
                    raw_data,
                    "error",
                    str(e)[:500],
                )

    return parsed_rows, event_count




def _process_manufacturing_workbook(db: Session, wb, upload_row: UploadFileModel, team_name: str) -> tuple[int, int]:
    source_type = "manufacturing"
    seen: set[tuple] = set()
    parsed_rows = 0
    event_count = 0
    group_counters: dict[tuple[str, str], int] = {}

    for ws in wb.worksheets:
        prepared = _prepare_sheet(ws, ["모델", "차분", "발주수량"])
        if not prepared:
            continue

        header_row, header_map, headers, _ = prepared

        model_col = _find_col(header_map, "모델")
        batch_col = _find_col(header_map, "차분", "st 차분")
        qty_col = _find_col(header_map, "발주수량")

        purchase_request_col = _find_col(header_map, "구매요청일자", "구매 요청일자")
        order_complete_col = _find_col(header_map, "발주완료일자", "발주 완료일자")
        ld_material_in_col = _find_col(header_map, "ld 자재입고", "ld자재입고")
        st_material_in_col = _find_col(header_map, "st 자재입고", "st자재입고")
        z_axis_col = _find_col(header_map, "z axis")
        ld_delivery_expected_col = _find_col(header_map, "ld 납품 예정일", "ld납품예정일")
        ld_delivery_done_col = _find_col(header_map, "ld 납품일", "ld납품일")
        inhouse_expected_col = _find_col(header_map, "사내 입고 예정일", "사내입고예정일")
        inhouse_done_col = _find_col(header_map, "사내 입고일", "사내입고일")

        st_vendor_col = _find_col(header_map, "st 외주처")
        ld_vendor_col = _find_col(header_map, "ld 외주처")
        year_col = _find_col(header_map, "년도")
        month_col = _find_col(header_map, "월")
        remark_col = _find_col(header_map, "비고")

        if not model_col or not batch_col or not qty_col:
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_data = _collect_raw_data(ws, row_idx, headers)
            if not raw_data:
                continue

            parsed_rows += 1

            try:
                model_text = _normalize_model_text(ws.cell(row_idx, model_col).value) if model_col else None
                batch_text = _normalize_batch_text(ws.cell(row_idx, batch_col).value) if batch_col else None
                qty = _parse_positive_int(ws.cell(row_idx, qty_col).value) if qty_col else None

                if not model_text or not batch_text or not qty:
                    _save_raw_row(
                        db,
                        upload_row.id,
                        ws.title,
                        row_idx,
                        None,
                        raw_data,
                        "parsed",
                        None,
                    )
                    continue

                group_key = (model_text, batch_text)
                start_unit = group_counters.get(group_key, 1)
                end_unit = start_unit + qty - 1
                representative_machine = _build_generated_machine_no(model_text, batch_text, start_unit)

                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    representative_machine,
                    raw_data,
                    "parsed",
                    None,
                )

                mo_no = f"{model_text}-{batch_text}"

                common_extra: dict[str, Any] = {
                    "generated_by": "manufacturing_qty_expand",
                    "manufacturing_model": model_text,
                    "manufacturing_batch": batch_text,
                    "order_qty": qty,
                    "unit_range_start": start_unit,
                    "unit_range_end": end_unit,
                }

                st_vendor = _text(ws.cell(row_idx, st_vendor_col).value) if st_vendor_col else None
                ld_vendor = _text(ws.cell(row_idx, ld_vendor_col).value) if ld_vendor_col else None
                year_value = ws.cell(row_idx, year_col).value if year_col else None
                month_value = ws.cell(row_idx, month_col).value if month_col else None
                remark_value = _text(ws.cell(row_idx, remark_col).value) if remark_col else None

                if st_vendor:
                    common_extra["st_vendor"] = st_vendor
                if ld_vendor:
                    common_extra["ld_vendor"] = ld_vendor
                if year_value is not None:
                    common_extra["plan_year"] = _jsonable(year_value)
                if month_value is not None:
                    common_extra["plan_month"] = _jsonable(month_value)
                if remark_value:
                    common_extra["remark"] = remark_value

                for unit_no in range(start_unit, end_unit + 1):
                    machine_no = _build_generated_machine_no(model_text, batch_text, unit_no)

                    equipment = _upsert_equipment(
                        db,
                        machine_no=machine_no,
                        model=model_text,
                        current_status="제조일정",
                        is_shipped=False,
                    )

                    unit_extra = dict(common_extra)
                    unit_extra["unit_no"] = unit_no
                    unit_extra["generated_machine_no"] = machine_no

                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="manufacturing_purchase_request",
                        event_name="구매요청",
                        event_date=_to_date(ws.cell(row_idx, purchase_request_col).value) if purchase_request_col else None,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data=unit_extra,
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="manufacturing_order_complete",
                        event_name="발주완료",
                        event_date=_to_date(ws.cell(row_idx, order_complete_col).value) if order_complete_col else None,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data=unit_extra,
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="manufacturing_ld_material_in",
                        event_name="LD 자재입고",
                        event_date=_to_date(ws.cell(row_idx, ld_material_in_col).value) if ld_material_in_col else None,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data=unit_extra,
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="manufacturing_st_material_in",
                        event_name="ST 자재입고",
                        event_date=_to_date(ws.cell(row_idx, st_material_in_col).value) if st_material_in_col else None,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data=unit_extra,
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="manufacturing_z_axis",
                        event_name="Z Axis",
                        event_date=_to_date(ws.cell(row_idx, z_axis_col).value) if z_axis_col else None,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data=unit_extra,
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="manufacturing_ld_delivery_expected",
                        event_name="LD 납품예정",
                        event_date=_to_date(ws.cell(row_idx, ld_delivery_expected_col).value) if ld_delivery_expected_col else None,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data=unit_extra,
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="manufacturing_ld_delivery_done",
                        event_name="LD 납품",
                        event_date=_to_date(ws.cell(row_idx, ld_delivery_done_col).value) if ld_delivery_done_col else None,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data=unit_extra,
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="manufacturing_inhouse_expected",
                        event_name="사내 입고예정",
                        event_date=_to_date(ws.cell(row_idx, inhouse_expected_col).value) if inhouse_expected_col else None,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data=unit_extra,
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                    event_count += _add_event(
                        db,
                        seen,
                        equipment_id=equipment.id,
                        source_type=source_type,
                        event_type="manufacturing_inhouse_done",
                        event_name="사내 입고",
                        event_date=_to_date(ws.cell(row_idx, inhouse_done_col).value) if inhouse_done_col else None,
                        status=None,
                        team_name=team_name,
                        mo_no=mo_no,
                        extra_data=unit_extra,
                        source_file_id=upload_row.id,
                        source_sheet_name=ws.title,
                        source_row_no=row_idx,
                    )

                group_counters[group_key] = end_unit + 1

            except Exception as e:
                _save_raw_row(
                    db,
                    upload_row.id,
                    ws.title,
                    row_idx,
                    None,
                    raw_data,
                    "error",
                    str(e)[:500],
                )

    return parsed_rows, event_count



# ------------------------------------------------------------------
# 조회 API
# ------------------------------------------------------------------
@router.get("/equipment")
def list_equipment(
    search: str = "",
    tab: str = "pending",
    page: int = 1,
    page_size: int = 50,
    db: Session = Depends(get_db),
):
    debug_db = db.execute(text("select current_database()")).scalar()
    debug_schema = db.execute(text("select current_schema()")).scalar()
    debug_search_path = db.execute(text("show search_path")).scalar()
    debug_equipment_count = db.execute(text("select count(*) from equipment_master")).scalar()
    debug_upload_count = db.execute(text("select count(*) from upload_files")).scalar()
    debug_raw_count = db.execute(text("select count(*) from import_raw_rows")).scalar()
    debug_event_count = db.execute(text("select count(*) from schedule_events")).scalar()

    print("DEBUG current_database =", debug_db)
    print("DEBUG current_schema =", debug_schema)
    print("DEBUG search_path =", debug_search_path)
    print("DEBUG equipment_master count =", debug_equipment_count)
    print("DEBUG upload_files count =", debug_upload_count)
    print("DEBUG import_raw_rows count =", debug_raw_count)
    print("DEBUG schedule_events count =", debug_event_count)

    tab = (tab or "pending").strip().lower()
    if tab not in ("pending", "shipped", "all"):
        tab = "pending"

    last_event_date_sq = (
        db.query(func.max(ScheduleEvent.event_date))
        .filter(ScheduleEvent.equipment_id == EquipmentMaster.id)
        .correlate(EquipmentMaster)
        .scalar_subquery()
    )

    last_event_name_sq = (
        db.query(ScheduleEvent.event_name)
        .filter(ScheduleEvent.equipment_id == EquipmentMaster.id)
        .order_by(ScheduleEvent.event_date.desc(), ScheduleEvent.id.desc())
        .limit(1)
        .correlate(EquipmentMaster)
        .scalar_subquery()
    )

    q = db.query(
        EquipmentMaster.id.label("id"),
        EquipmentMaster.machine_no.label("machine_no"),
        EquipmentMaster.model.label("model"),
        EquipmentMaster.customer_name.label("customer_name"),
        EquipmentMaster.cold_type.label("cold_type"),
        EquipmentMaster.current_status.label("current_status"),
        EquipmentMaster.is_shipped.label("is_shipped"),
        last_event_name_sq.label("last_event_name"),
        last_event_date_sq.label("last_event_date"),
    )

    keyword = (search or "").strip()
    if keyword:
        like = f"%{keyword}%"
        q = q.filter(
            or_(
                EquipmentMaster.machine_no.ilike(like),
                EquipmentMaster.model.ilike(like),
                EquipmentMaster.customer_name.ilike(like),
            )
        )

    if tab == "pending":
        q = q.filter(EquipmentMaster.is_shipped.is_(False))
    elif tab == "shipped":
        q = q.filter(EquipmentMaster.is_shipped.is_(True))

    total = q.count()

    rows = (
        q.order_by(EquipmentMaster.is_shipped.asc(), EquipmentMaster.machine_no.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "items": [
            {
                "id": row.id,
                "machine_no": row.machine_no,
                "model": row.model,
                "customer_name": row.customer_name,
                "cold_type": row.cold_type,
                "current_status": row.current_status,
                "is_shipped": bool(row.is_shipped),
                "last_event_name": row.last_event_name,
                "last_event_date": row.last_event_date,
            }
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/equipment/{equipment_id}")
def get_equipment_detail(
    equipment_id: int,
    db: Session = Depends(get_db),
):
    equipment = (
        db.query(EquipmentMaster)
        .filter(EquipmentMaster.id == equipment_id)
        .first()
    )
    if not equipment:
        raise HTTPException(status_code=404, detail="장비를 찾을 수 없습니다.")

    rows = (
        db.query(ScheduleEvent)
        .filter(ScheduleEvent.equipment_id == equipment_id)
        .order_by(ScheduleEvent.event_date.asc(), ScheduleEvent.id.asc())
        .all()
    )

    events = []
    changed_count = 0

    for row in rows:
        extra = row.extra_data or {}
        previous_date = extra.get("previous_date")
        is_changed = (row.status == "변경") or bool(previous_date)

        if is_changed:
            changed_count += 1

        events.append(
            {
                "id": row.id,
                "source_type": row.source_type,
                "event_type": row.event_type,
                "event_name": row.event_name,
                "event_date": row.event_date,
                "status": row.status,
                "team_name": row.team_name,
                "mo_no": row.mo_no,
                "previous_date": previous_date,
                "is_changed": is_changed,
                "extra_data": extra,
            }
        )

    return {
        "equipment": {
            "id": equipment.id,
            "machine_no": equipment.machine_no,
            "model": equipment.model,
            "customer_name": equipment.customer_name,
            "stage_sn": equipment.stage_sn,
            "loader_sn": equipment.loader_sn,
            "cold_type": equipment.cold_type,
            "mani_type": equipment.mani_type,
            "current_status": equipment.current_status,
            "is_shipped": bool(equipment.is_shipped),
            "created_at": equipment.created_at,
            "updated_at": equipment.updated_at,
        },
        "events": events,
        "changed_count": changed_count,
    }




def _parse_batch_query(query: str) -> tuple[str, str]:
    raw = re.sub(r"\s+", "", (query or ""))
    if not raw:
        raise HTTPException(status_code=400, detail="검색어를 입력해주세요. 예: D(e)-12")

    m = re.fullmatch(r"(.+)-(\d{1,2})", raw)
    if not m:
        raise HTTPException(
            status_code=400,
            detail="검색 형식이 올바르지 않습니다. 예: D(e)-12",
        )

    model = m.group(1)
    batch = m.group(2).zfill(2)
    return model, batch


def _extract_model_batch_from_machine_no(machine_no: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    s = _text(machine_no)
    if not s:
        return None, None
    s = re.sub(r"\s+", "", s)

    m = re.fullmatch(r"(.+)-(\d{2})-(\d{2,})", s)
    if m:
        return m.group(1), m.group(2)

    m = re.fullmatch(r"(.+)-(\d{2})", s)
    if m:
        return m.group(1), m.group(2)

    return None, None


@router.get("/batch-history")
def get_batch_history(
    query: str,
    db: Session = Depends(get_db),
):
    model, batch = _parse_batch_query(query)
    like_pattern = f"{model}-{batch}-%"

    equipment_rows = (
        db.query(EquipmentMaster)
        .filter(EquipmentMaster.machine_no.like(like_pattern))
        .order_by(EquipmentMaster.machine_no.asc())
        .all()
    )

    if not equipment_rows:
        raise HTTPException(status_code=404, detail=f"{model}-{batch} 차분 장비를 찾을 수 없습니다.")

    equipment_ids = [int(row.id) for row in equipment_rows]
    machine_map = {int(row.id): row.machine_no for row in equipment_rows}

    current_rows = (
        db.query(ScheduleEvent)
        .filter(ScheduleEvent.equipment_id.in_(equipment_ids))
        .order_by(
            ScheduleEvent.event_date.asc(),
            ScheduleEvent.event_name.asc(),
            ScheduleEvent.id.asc(),
        )
        .all()
    )

    history_rows = (
        db.query(ScheduleEventHistory)
        .filter(ScheduleEventHistory.equipment_id.in_(equipment_ids))
        .order_by(
            ScheduleEventHistory.created_at.desc(),
            ScheduleEventHistory.id.desc(),
        )
        .all()
    )

    current_summary_map: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in current_rows:
        key = (row.source_type, row.event_type, row.event_name)
        item = current_summary_map.get(key)
        if not item:
            item = {
                "source_type": row.source_type,
                "event_type": row.event_type,
                "event_name": row.event_name,
                "machine_count": 0,
                "min_event_date": row.event_date,
                "max_event_date": row.event_date,
            }
            current_summary_map[key] = item

        item["machine_count"] += 1
        if row.event_date and (item["min_event_date"] is None or row.event_date < item["min_event_date"]):
            item["min_event_date"] = row.event_date
        if row.event_date and (item["max_event_date"] is None or row.event_date > item["max_event_date"]):
            item["max_event_date"] = row.event_date

    inserted_count = 0
    updated_count = 0
    deleted_count = 0
    latest_changed_at = None
    history_event_name_map: dict[str, int] = {}

    history_items = []
    for row in history_rows:
        if row.change_type == "inserted":
            inserted_count += 1
        elif row.change_type == "deleted":
            deleted_count += 1
        else:
            updated_count += 1

        history_event_name_map[row.event_name] = history_event_name_map.get(row.event_name, 0) + 1

        if row.created_at and (latest_changed_at is None or row.created_at > latest_changed_at):
            latest_changed_at = row.created_at

        history_items.append(
            {
                "id": row.id,
                "equipment_id": int(row.equipment_id),
                "machine_no": machine_map.get(int(row.equipment_id)),
                "source_type": row.source_type,
                "event_type": row.event_type,
                "event_name": row.event_name,
                "team_name": row.team_name,
                "mo_no": row.mo_no,
                "change_type": row.change_type,
                "before_event_date": row.before_event_date,
                "before_status": row.before_status,
                "before_extra_data": row.before_extra_data or {},
                "after_event_date": row.after_event_date,
                "after_status": row.after_status,
                "after_extra_data": row.after_extra_data or {},
                "changed_by": row.changed_by,
                "change_reason": row.change_reason,
                "created_at": row.created_at,
            }
        )

    current_items = []
    for row in current_rows:
        current_items.append(
            {
                "id": row.id,
                "equipment_id": int(row.equipment_id),
                "machine_no": machine_map.get(int(row.equipment_id)),
                "source_type": row.source_type,
                "event_type": row.event_type,
                "event_name": row.event_name,
                "event_date": row.event_date,
                "status": row.status,
                "team_name": row.team_name,
                "mo_no": row.mo_no,
                "extra_data": row.extra_data or {},
            }
        )

    current_summary = sorted(
        current_summary_map.values(),
        key=lambda x: (
            x.get("min_event_date") or date.max,
            x.get("event_name") or "",
        ),
    )
    history_event_summary = [
        {"event_name": event_name, "change_count": count}
        for event_name, count in sorted(
            history_event_name_map.items(),
            key=lambda x: (-x[1], x[0]),
        )
    ]

    return {
        "query": query,
        "model": model,
        "batch": batch,
        "machine_count": len(equipment_rows),
        "machines": [row.machine_no for row in equipment_rows],
        "history_summary": {
            "total_changes": len(history_rows),
            "inserted_count": inserted_count,
            "updated_count": updated_count,
            "deleted_count": deleted_count,
            "latest_changed_at": latest_changed_at,
        },
        "current_summary": current_summary,
        "history_event_summary": history_event_summary,
        "current_events": current_items,
        "history_events": history_items,
    }


@router.get("/upload-history")
def get_upload_history(db: Session = Depends(get_db)):
    rows = (
        db.query(UploadFileModel)
        .order_by(UploadFileModel.created_at.desc(), UploadFileModel.id.desc())
        .limit(100)
        .all()
    )

    return [
        {
            "id": row.id,
            "team_name": row.team_name,
            "file_name": row.file_name,
            "uploaded_by": row.uploaded_by,
            "upload_status": row.upload_status,
            "message": row.message,
            "created_at": row.created_at,
        }
        for row in rows
    ]


# ------------------------------------------------------------------
# 업로드 API (A 방식)
# ------------------------------------------------------------------
@router.post("/upload")
async def upload_schedule_excel(
    team_name: str = Form(...),
    uploaded_by: str = Form(""),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    source_type = _normalize_team_name(team_name)

    if not file.filename:
        raise HTTPException(status_code=400, detail="파일명이 없습니다.")

    lower_name = file.filename.lower()
    if not (lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="xlsx 또는 xlsm 파일만 업로드 가능합니다.")

    save_name = f"{uuid.uuid4().hex}_{file.filename}"
    save_path = UPLOAD_DIR / save_name

    content = await file.read()
    save_path.write_bytes(content)

    upload_row = UploadFileModel(
        team_name=team_name,
        file_name=file.filename,
        file_path=str(save_path),
        uploaded_by=(uploaded_by or "").strip() or None,
        upload_status="uploaded",
        message="업로드 완료",
    )
    db.add(upload_row)
    db.commit()
    db.refresh(upload_row)

    try:
        wb = load_workbook(save_path, data_only=True)

        parsed_rows = 0
        event_count = 0

        if source_type == "production":
            parsed_rows, event_count = _process_production_workbook(db, wb, upload_row, team_name)
        elif source_type == "shipment":
            parsed_rows, event_count = _process_shipment_workbook(db, wb, upload_row, team_name)
        elif source_type == "remodel":
            parsed_rows, event_count = _process_remodel_workbook(db, wb, upload_row, team_name)
        elif source_type == "interface":
            parsed_rows, event_count = _process_interface_workbook(db, wb, upload_row, team_name)
        elif source_type == "mani":
            parsed_rows, event_count = _process_mani_workbook(db, wb, upload_row, team_name)
        elif source_type == "opus":
            parsed_rows, event_count = _process_opus_workbook(db, wb, upload_row, team_name)
        elif source_type == "chiller":
            parsed_rows, event_count = _process_chiller_workbook(db, wb, upload_row, team_name)
        elif source_type == "manufacturing":
            parsed_rows, event_count = _process_manufacturing_workbook(db, wb, upload_row, team_name)
        else:
            raise HTTPException(status_code=400, detail="지원하지 않는 team_name 입니다.")

        # 0행이면 성공 처리하지 말기
        if parsed_rows == 0:
            raise ValueError("엑셀에서 읽을 수 있는 시트/헤더를 찾지 못했습니다.")

        # OPUS 제외하고 이벤트가 0개면 경고성 실패 처리
        if source_type != "opus" and event_count == 0:
            raise ValueError("장비는 읽었지만 일정 날짜를 인식하지 못해 이벤트가 0건입니다.")

        inserted_count, updated_count, deleted_count = _record_schedule_event_history(
            db,
            source_type=source_type,
            new_source_file_id=upload_row.id,
            changed_by=upload_row.uploaded_by,
        )
        history_count = inserted_count + updated_count + deleted_count

        # 새 업로드 데이터는 살리고, 이전 source_type 데이터만 삭제
        db.query(ScheduleEvent).filter(
            ScheduleEvent.source_type == source_type,
            ScheduleEvent.source_file_id != upload_row.id,
        ).delete(synchronize_session=False)

        upload_row.upload_status = "success"
        upload_row.message = (
            f"{parsed_rows}행 처리 / {event_count}개 이벤트 반영 / "
            f"history {history_count}건 (신규 {inserted_count}, 변경 {updated_count}, 삭제 {deleted_count})"
        )
        db.commit()

        return {
            "ok": True,
            "upload_file_id": upload_row.id,
            "team_name": team_name,
            "source_type": source_type,
            "parsed_rows": parsed_rows,
            "event_count": event_count,
            "history_count": history_count,
            "history_inserted_count": inserted_count,
            "history_updated_count": updated_count,
            "history_deleted_count": deleted_count,
            "message": upload_row.message,
        }

    except Exception as e:
        db.rollback()

        fail_row = db.query(UploadFileModel).filter(UploadFileModel.id == upload_row.id).first()
        if fail_row:
            fail_row.upload_status = "fail"
            fail_row.message = str(e)[:500]
            db.commit()

        raise HTTPException(status_code=400, detail=f"엑셀 처리 실패: {e}")